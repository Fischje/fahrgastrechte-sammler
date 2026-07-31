from __future__ import annotations

import csv
import io
import os
import re
import secrets
import shutil
from datetime import date, datetime
from pathlib import Path

from flask import Flask, Response, flash, redirect, render_template, request, send_file, url_for
from flask_login import LoginManager, UserMixin, current_user, login_required, login_user, logout_user
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash


db = SQLAlchemy()
login_manager = LoginManager()
login_manager.login_view = "login"


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class Journey(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    journey_date = db.Column(db.Date, nullable=False, index=True)
    train_number = db.Column(db.String(60), nullable=False, index=True)
    departure_station = db.Column(db.String(160), nullable=False)
    destination_station = db.Column(db.String(160), nullable=False)
    scheduled_departure = db.Column(db.String(5), default="")
    scheduled_arrival = db.Column(db.String(5), default="")
    actual_departure = db.Column(db.String(5), default="")
    actual_arrival = db.Column(db.String(5), default="")
    delay_minutes = db.Column(db.Integer, default=0, nullable=False)
    cancelled = db.Column(db.Boolean, default=False, nullable=False)
    reason = db.Column(db.String(500), default="")
    ticket_price = db.Column(db.Float, default=0.0, nullable=False)
    compensation_claimed = db.Column(db.Boolean, default=False, nullable=False)
    compensation_received = db.Column(db.Float, default=0.0, nullable=False)
    notes = db.Column(db.Text, default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)


@login_manager.user_loader
def load_user(user_id: str):
    return db.session.get(User, int(user_id))


def password_error(password: str) -> str | None:
    if len(password) < 6:
        return "Das Passwort muss mindestens 6 Zeichen lang sein."
    if not re.search(r"[A-Za-zÄÖÜäöüß]", password):
        return "Das Passwort muss mindestens einen Buchstaben enthalten."
    if not re.search(r"\d", password):
        return "Das Passwort muss mindestens eine Zahl enthalten."
    if not re.search(r"[^A-Za-zÄÖÜäöüß0-9]", password):
        return "Das Passwort muss mindestens ein Sonderzeichen enthalten."
    return None


def parse_float(value: str) -> float:
    value = (value or "").strip().replace(".", "").replace(",", ".")
    try:
        return max(0.0, float(value))
    except ValueError:
        return 0.0


def parse_int(value: str) -> int:
    try:
        return max(0, int(value or 0))
    except ValueError:
        return 0


def ensure_secret(data_dir: Path) -> str:
    secret_file = data_dir / "secret_key"
    if not secret_file.exists():
        secret_file.write_text(secrets.token_urlsafe(64), encoding="utf-8")
        try:
            secret_file.chmod(0o600)
        except OSError:
            pass
    return secret_file.read_text(encoding="utf-8").strip()


def create_app() -> Flask:
    data_dir = Path(os.environ.get("APP_DATA_DIR", "/data"))
    backup_dir = Path(os.environ.get("APP_BACKUP_DIR", "/backups"))
    data_dir.mkdir(parents=True, exist_ok=True)
    backup_dir.mkdir(parents=True, exist_ok=True)

    app = Flask(__name__)
    app.config.update(
        SECRET_KEY=ensure_secret(data_dir),
        SQLALCHEMY_DATABASE_URI=f"sqlite:///{data_dir / 'fahrgastrechte.db'}",
        SQLALCHEMY_TRACK_MODIFICATIONS=False,
        MAX_CONTENT_LENGTH=10 * 1024 * 1024,
    )
    db.init_app(app)
    login_manager.init_app(app)

    with app.app_context():
        db.create_all()
        # Lightweight SQLite migration for updates from earlier versions.
        columns = {row[1] for row in db.session.execute(db.text("PRAGMA table_info(journey)")).all()}
        for name, definition in {
            "actual_departure": "VARCHAR(5) DEFAULT ''",
            "actual_arrival": "VARCHAR(5) DEFAULT ''",
        }.items():
            if name not in columns:
                db.session.execute(db.text(f"ALTER TABLE journey ADD COLUMN {name} {definition}"))
        db.session.commit()

    @app.get("/health")
    def health():
        return {"status": "ok", "version": "2.2.0", "app": "Fahrgastrechte-Sammler"}

    @app.before_request
    def require_setup():
        allowed = {"setup", "static", "health"}
        if request.endpoint in allowed or request.endpoint is None:
            return None
        if not User.query.first():
            return redirect(url_for("setup"))
        return None

    @app.route("/setup", methods=["GET", "POST"])
    def setup():
        if User.query.first():
            return redirect(url_for("login"))
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            confirmation = request.form.get("password_confirmation", "")
            if len(username) < 3:
                flash("Der Benutzername muss mindestens 3 Zeichen lang sein.", "danger")
            elif password != confirmation:
                flash("Die Passwörter stimmen nicht überein.", "danger")
            elif error := password_error(password):
                flash(error, "danger")
            else:
                user = User(username=username, password_hash=generate_password_hash(password))
                db.session.add(user)
                db.session.commit()
                login_user(user)
                flash("Die Ersteinrichtung wurde abgeschlossen.", "success")
                return redirect(url_for("index"))
        return render_template("setup.html")

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for("index"))
        if request.method == "POST":
            user = User.query.filter_by(username=request.form.get("username", "").strip()).first()
            if user and check_password_hash(user.password_hash, request.form.get("password", "")):
                login_user(user)
                return redirect(url_for("index"))
            flash("Benutzername oder Passwort ist falsch.", "danger")
        return render_template("login.html")

    @app.post("/logout")
    @login_required
    def logout():
        logout_user()
        return redirect(url_for("login"))

    @app.get("/")
    @login_required
    def index():
        selected_year = request.args.get("year", type=int) or date.today().year
        journeys = Journey.query.filter(db.extract("year", Journey.journey_date) == selected_year).order_by(Journey.journey_date.desc(), Journey.id.desc()).all()
        years = [row[0] for row in db.session.query(db.extract("year", Journey.journey_date)).distinct().order_by(db.extract("year", Journey.journey_date).desc()).all()]
        years = [int(y) for y in years if y]
        if date.today().year not in years:
            years.insert(0, date.today().year)
        stats = {
            "count": len(journeys),
            "cancelled": sum(1 for j in journeys if j.cancelled),
            "delay": sum(j.delay_minutes or 0 for j in journeys),
            "claimed": sum(1 for j in journeys if j.compensation_claimed),
        }
        return render_template("index.html", journeys=journeys, years=years, selected_year=selected_year, stats=stats)

    def fill_journey(journey: Journey):
        journey.journey_date = datetime.strptime(request.form["journey_date"], "%Y-%m-%d").date()
        journey.train_number = request.form.get("train_number", "").strip()
        journey.departure_station = request.form.get("departure_station", "").strip()
        journey.destination_station = request.form.get("destination_station", "").strip()
        journey.scheduled_departure = request.form.get("scheduled_departure", "")
        journey.scheduled_arrival = request.form.get("scheduled_arrival", "")
        journey.actual_departure = request.form.get("actual_departure", "")
        journey.actual_arrival = request.form.get("actual_arrival", "")
        journey.delay_minutes = parse_int(request.form.get("delay_minutes", "0"))
        journey.cancelled = request.form.get("cancelled") == "on"
        journey.reason = request.form.get("reason", "").strip()
        journey.compensation_claimed = request.form.get("compensation_claimed") == "on"
        journey.notes = request.form.get("notes", "").strip()

    def distinct_values(column):
        return [row[0] for row in db.session.query(column).filter(column != "").distinct().order_by(column).all() if row[0]]

    def get_suggestions():
        stations = sorted(set(distinct_values(Journey.departure_station) + distinct_values(Journey.destination_station)), key=str.casefold)
        return {
            "train_numbers": distinct_values(Journey.train_number),
            "stations": stations,
            "reasons": distinct_values(Journey.reason),
        }

    @app.route("/journey/new", methods=["GET", "POST"])
    @login_required
    def journey_new():
        if request.method == "POST":
            journey = Journey()
            try:
                fill_journey(journey)
                if not journey.train_number or not journey.departure_station or not journey.destination_station:
                    raise ValueError("Pflichtfelder fehlen")
                db.session.add(journey)
                db.session.commit()
                flash("Die Fahrt wurde gespeichert.", "success")
                return redirect(url_for("index", year=journey.journey_date.year))
            except (ValueError, KeyError):
                flash("Bitte fülle Datum, Zugnummer, Start und Ziel vollständig aus.", "danger")
        suggestions = get_suggestions()
        return render_template("journey_form.html", journey=None, today=date.today().isoformat(), **suggestions)

    @app.route("/journey/<int:journey_id>/edit", methods=["GET", "POST"])
    @login_required
    def journey_edit(journey_id: int):
        journey = db.get_or_404(Journey, journey_id)
        if request.method == "POST":
            try:
                fill_journey(journey)
                db.session.commit()
                flash("Die Fahrt wurde aktualisiert.", "success")
                return redirect(url_for("index", year=journey.journey_date.year))
            except (ValueError, KeyError):
                flash("Die Eingaben konnten nicht gespeichert werden.", "danger")
        suggestions = get_suggestions()
        return render_template("journey_form.html", journey=journey, today=date.today().isoformat(), **suggestions)

    @app.post("/journeys/bulk")
    @login_required
    def journeys_bulk():
        selected_year = request.form.get("year", type=int) or date.today().year
        action = request.form.get("bulk_action", "").strip()
        raw_ids = request.form.getlist("journey_ids")
        try:
            journey_ids = sorted({int(value) for value in raw_ids})
        except ValueError:
            journey_ids = []

        if not journey_ids:
            flash("Bitte wähle mindestens einen Eintrag aus.", "warning")
            return redirect(url_for("index", year=selected_year))

        journeys = Journey.query.filter(Journey.id.in_(journey_ids)).all()
        if action == "delete":
            count = len(journeys)
            for journey in journeys:
                db.session.delete(journey)
            db.session.commit()
            flash(f"{count} Einträge wurden gelöscht.", "warning")
        elif action == "claim":
            count = 0
            for journey in journeys:
                if not journey.compensation_claimed:
                    journey.compensation_claimed = True
                    count += 1
            db.session.commit()
            flash(f"{count} Einträge wurden als Entschädigung eingereicht markiert.", "success")
        else:
            flash("Bitte wähle eine gültige Sammelaktion aus.", "danger")

        return redirect(url_for("index", year=selected_year))

    @app.post("/journey/<int:journey_id>/delete")
    @login_required
    def journey_delete(journey_id: int):
        journey = db.get_or_404(Journey, journey_id)
        year = journey.journey_date.year
        db.session.delete(journey)
        db.session.commit()
        flash("Der Eintrag wurde gelöscht.", "warning")
        return redirect(url_for("index", year=year))

    def export_rows(year: int):
        return Journey.query.filter(db.extract("year", Journey.journey_date) == year).order_by(Journey.journey_date).all()

    headers = ["Datum", "Zugnummer", "Start", "Ziel", "Plan-Abfahrt", "Tatsächliche Abfahrt", "Plan-Ankunft", "Tatsächliche Ankunft", "Verspätung (Min.)", "Ausfall", "Grund / Störung", "Notizen", "Entschädigung beantragt"]

    @app.get("/export/csv/<int:year>")
    @login_required
    def export_csv(year: int):
        output = io.StringIO()
        output.write("\ufeff")
        writer = csv.writer(output, delimiter=";", lineterminator="\n")
        writer.writerow(headers)
        for j in export_rows(year):
            writer.writerow([j.journey_date.strftime("%d.%m.%Y"), j.train_number, j.departure_station, j.destination_station, j.scheduled_departure, j.actual_departure, j.scheduled_arrival, j.actual_arrival, j.delay_minutes, "Ja" if j.cancelled else "Nein", j.reason, j.notes, "Ja" if j.compensation_claimed else "Nein"])
        return Response(output.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": f"attachment; filename=fahrgastrechte-{year}.csv"})

    @app.get("/export/xlsx/<int:year>")
    @login_required
    def export_xlsx(year: int):
        wb = Workbook()
        ws = wb.active
        ws.title = str(year)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="245C4A")
            cell.alignment = Alignment(vertical="center")
        for j in export_rows(year):
            ws.append([j.journey_date, j.train_number, j.departure_station, j.destination_station, j.scheduled_departure, j.actual_departure, j.scheduled_arrival, j.actual_arrival, j.delay_minutes, "Ja" if j.cancelled else "Nein", j.reason, j.notes, "Ja" if j.compensation_claimed else "Nein"])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        widths = [13, 14, 24, 24, 15, 18, 15, 18, 18, 10, 32, 40, 24]
        from openpyxl.utils import get_column_letter
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return send_file(stream, as_attachment=True, download_name=f"fahrgastrechte-{year}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.post("/backup")
    @login_required
    def backup():
        source = data_dir / "fahrgastrechte.db"
        target = backup_dir / f"fahrgastrechte-{datetime.now().strftime('%Y%m%d-%H%M%S')}.db"
        db.session.remove()
        shutil.copy2(source, target)
        flash(f"Backup erstellt: {target.name}", "success")
        return redirect(url_for("index"))

    @app.route("/account", methods=["GET", "POST"])
    @login_required
    def account():
        if request.method == "POST":
            current_password = request.form.get("current_password", "")
            new_password = request.form.get("new_password", "")
            confirmation = request.form.get("password_confirmation", "")
            if not check_password_hash(current_user.password_hash, current_password):
                flash("Das aktuelle Passwort ist falsch.", "danger")
            elif new_password != confirmation:
                flash("Die neuen Passwörter stimmen nicht überein.", "danger")
            elif error := password_error(new_password):
                flash(error, "danger")
            else:
                current_user.password_hash = generate_password_hash(new_password)
                db.session.commit()
                flash("Das Passwort wurde geändert.", "success")
                return redirect(url_for("account"))
        return render_template("account.html")

    return app
